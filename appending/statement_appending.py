import fitz
import re
import json
import numpy as np
import cv2
import pytesseract
import pandas as pd
import datetime
from word2number import w2n
import os
from openpyxl import load_workbook
from PIL import ImageDraw, Image, ImageFont

new_height = 3500
new_width = 4500


def xlstoxlsx(input_directory, output_directory):
    xls_data = pd.read_excel(input_directory, sheet_name=None)
    with pd.ExcelWriter(output_directory) as writer:
        for sheet_name, df in xls_data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # return writer.book


def excelToImage(file_name):
    wb = load_workbook(file_name, data_only=True)
    image_list = []
    count = 1

    for sheets in wb.sheetnames:
        sheet = wb[sheets]

        max_row = sheet.max_row
        max_col = sheet.max_column
        cell_width = 3000
        cell_height = 100
        padding_x = 130
        padding_y = 50

        img = Image.new('RGB', ((max_col + 1) * (cell_width + padding_x), (max_row + 1) * (cell_height + padding_y)),
                        color='white')

        draw = ImageDraw.Draw(img)
        font = ImageFont.truetype("arialbd.ttf", size=120, encoding="unic")
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                if cell.input:
                    cell_input = str(cell.input)
                    x_pos = (cell.column - 1) * (cell_width + padding_x)
                    y_pos = (cell.row - 1) * (cell_height + padding_y)

                    draw.text((x_pos, y_pos), cell_input, fill='black', font=font)

        resized_img = img.resize((new_width, new_height))

        img_array = np.array(resized_img)
        image_list.append(img_array)
        # Save the resized image
        resized_img.save(f'excel_images/{count}.jpg')
        count += 1

    return image_list


def enhance_image(image_array, alpha=1.8, beta=40):
    enhanced_image = cv2.convertScaleAbs(image_array, alpha=alpha, beta=beta)
    return enhanced_image


def pdf_to_jpg(pdf_path):
    # Open the PDF file
    pdf_document = fitz.open(pdf_path, )
    images = []
    # arrays = []
    # jpg_folder = output_folder

    # Iterate through each page in the PDF
    if len(pdf_document) > 1:
        for page_number in range(len(pdf_document)):
            page = pdf_document[page_number]

            # Create an image of the page (DPI setting can be adjusted)
            px = page.get_pixmap(matrix=fitz.Matrix(300 / 72, 300 / 72))

            # Convert Pixmap object to Numpy array
            image_array = np.frombuffer(
                px.samples, dtype=np.uint8).reshape(px.h, px.w, px.n)

            enhanced_image = enhance_image(image_array)

            images.append(enhanced_image)

            # Save the image using OpenCV
            # jpg_path = os.path.join(output_folder, f"page_{page_number + 1}.jpg")
            # cv2.imwrite(jpg_path, cv2.cvtColor(image_array, cv2.COLOR_RGB2BGR))
    else:
        page = pdf_document[len(pdf_document)]
        px = page.get_pixmap()
        image_array = np.frombuffer(
            px.samples, dtype=np.uint8).reshape(px.h, px.w, px.n)
        images.append(image_array)

        # jpg_path = os.path.join(output_folder, "page_1.jpg")
        # cv2.imwrite(jpg_path, cv2.cvtColor(image_array, cv2.COLOR_RGB2BGR))

    # Close the PDF file
    pdf_document.close()
    # np.save(os.path.join(jpg_folder, "arrays.npy"), np.array(arrays))
    return images


def extract_years_from_dates(date_list):
    print("Date list to extraction function:", date_list)
    years = set()
    year_pattern = re.compile(r'\b\d{4}\b')

    current_year = datetime.datetime.today().year

    for date_str in date_list:
        year_matches = re.findall(year_pattern, date_str)
        for year in year_matches:
            year_int = int(year)
            if current_year - 5 <= year_int <= current_year:
                years.add(year)

    return list(years)


def image_to_text(OCR_path, *args):
    master_list = []
    for arg in args:
        if not isinstance(arg, list):
            raise TypeError("Input Parameter must be a list of numpy arrays")
        else:
            pytesseract.pytesseract.tesseract_cmd = OCR_path
            final_json = {}

            text = ""
            for _, images in enumerate(arg[267:268]):
                gray_images = cv2.cvtColor(images, cv2.COLOR_BGR2GRAY)
                threshold_img = cv2.adaptiveThreshold(gray_images, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 81, 15)
                extracted_text = pytesseract.image_to_string(threshold_img, config=r'--psm 4')
                print("Extracted Text:", extracted_text)
                text = text + "\n" + extracted_text
            print("Text:", text)
            lines = text.split("\n")
            print("Lines:", lines)
            date_pattern = r'\b(?:\d{1,2}[-/]\d{1,2}[-/]\d{2,4}|' \
                           r'\d{1,2} (?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]* \d{2,4}|' \
                           r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]* \d{1,2})\b'

            date_pattern2 = r'\b(?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|' \
                            r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-zA-Z.,-]*[\s-]?\d{1,2}?[,\s-]?[\s]?\d{4}|' \
                            r'\d{1,2}[/-]\d{4}|\d{4})\b(?!-?\d{2,4})' \
                            r"\b(?:January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s*\d{4}\b"

            date_pattern3 = (r"\b(?:January|February|March|April|May|June|July|August|September|October|November"
                             r"|December)\s+\d{1,2},\s*\d{4}\b")

            dates = []
            for inputs in lines:
                if re.findall(date_pattern, inputs):
                    dates.append(inputs)
            print("All date related values in lines:", dates)
            supp_dates = []
            for date in dates:
                if bool(re.search(r'[a-zA-z]', date)) and bool(re.search(r'\d', date)):
                    supp_dates.append(date)
            print("ALl supp_date values:", supp_dates)

            def keep_int_or_float(data_check):
                try:
                    if isinstance(int(data_check), int):
                        return int(data_check)
                except:
                    try:
                        if isinstance(float(data_check), float):
                            return float(data_check)
                    except:
                        return None

            num_pattern = r'[-+]?\d{1,5}(?:,?\d{3})*(?:\.\d+)?|\(\s*[-+]?\d{1,3}(?:,?\d{3})*(?:\.\d+)?\s*\)'
            pattern = r'[,:()\-]'
            year_list = None
            print("Year List before anything:", year_list)
            unstruct_data = []
            for word_list in lines:
                print("Word List:", word_list)
                num_list = re.findall(num_pattern, word_list)
                num_list = [num.replace(',', '') for num in num_list]
                num_list = [
                    str(float(number.replace('(', '').replace(')', '')) * (-1)) if '(' in number else str(number)
                    for number in num_list]
                print("Num List:", num_list)
                word_list = re.sub(pattern, '', word_list)
                word_list = word_list.split()
                label = "_".join([word for word in word_list if word.isalpha()]).lower()
                print("Label:", label)
                if len(num_list) != 0:
                    num_list = list(map(keep_int_or_float, num_list))
                    print("Num List after mapping data type:", num_list)

                    if all(isinstance(x, int) for x in num_list):
                        if all((datetime.datetime.today().year - 5 <= x <= datetime.datetime.today().year) for x in num_list):
                            year_list = num_list
                        print("Year List 1:", year_list)
                    elif year_list is None:
                        dates3 = []
                        for elements in lines:
                            if re.findall(date_pattern3, elements):
                                dates3.append(elements)
                        year_list = extract_years_from_dates(dates3)
                        print("Year Lis from Date Pattern3:", year_list)

                    unstruct_data.append({label: num_list})

            print("Year List:", year_list)
            print("Unstruct Data:", unstruct_data)

            # dummy = []
            # if all(isinstance(x, str) for x in year_list):
            #     dummy.extend(int(re.findall(r'\d+', year)[0]) for year in year_list)
            #     year_list = sorted(dummy, reverse = True)
            # if year_list != None:
            #     for dictionaries in unstruct_data:
            #         for key in dictionaries.keys():
            #             if key != '':
            #                 length = len(dictionaries[key])
            #                 while length < len(year_list):
            #                     rem = len(year_list) - length
            #                     for i in range(rem):
            #                         dictionaries[key].insert(i,0)
            #                     length = len(dictionaries[key])
            #     for idx in range(len(year_list)):
            #         temp_list = []
            #         for dic in unstruct_data:
            #             for key, value in dic.items():
            #                 if key != '':
            #                     if len(value) > len(year_list):
            #                         s = len(value) - len(year_list) + idx
            #                         temp_list.append({key: value[s]})
            #                     else:
            #                         temp_list.append({key: value[idx]})
            #         final_json[year_list[idx]] = temp_list
            # master_list.append(final_json)
    # print(master_list)
    # return master_list


pdf_file = "zomato.pdf"  # Replace with your PDF file path
path_ocr = r"C:\Program Files\Tesseract-OCR\tesseract.exe"  # Replace with your path
output_folder = r"C:\Users\DELL\Desktop\pdf_extractor\jpg_folder"
img = pdf_to_jpg(pdf_file)

line_items = (image_to_text(path_ocr, img))
print("Line items extracted:", line_items)

with open('line_items_ambuja_pnl.json', 'w') as json_file:
    json.dump(line_items, json_file)
