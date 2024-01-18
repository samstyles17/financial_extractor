import fitz
import re
import json
import numpy as np
import cv2
import pytesseract
import pandas as pd
import datetime
from word2number import w2n
import os
from openpyxl import load_workbook
from PIL import ImageDraw, Image, ImageFont

new_height = 3500
new_width = 4500


def xlstoxlsx(input_directory, output_directory):
    xls_data = pd.read_excel(input_directory, sheet_name=None)
    with pd.ExcelWriter(output_directory) as writer:
        for sheet_name, df in xls_data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # return writer.book


def excelToImage(file_name):
    wb = load_workbook(file_name, data_only=True)
    image_list = []
    count = 1

    for sheets in wb.sheetnames:
        sheet = wb[sheets]

        max_row = sheet.max_row
        max_col = sheet.max_column
        cell_width = 3000
        cell_height = 100
        padding_x = 130
        padding_y = 50

        img = Image.new('RGB', ((max_col + 1) * (cell_width + padding_x), (max_row + 1) * (cell_height + padding_y)),
                        color='white')

        draw = ImageDraw.Draw(img)
        font = ImageFont.truetype("arialbd.ttf", size=120, encoding="unic")
        for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                if cell.value:
                    cell_value = str(cell.value)
                    x_pos = (cell.column - 1) * (cell_width + padding_x)
                    y_pos = (cell.row - 1) * (cell_height + padding_y)

                    draw.text((x_pos, y_pos), cell_value, fill='black', font=font)

        resized_img = img.resize((new_width, new_height))

        img_array = np.array(resized_img)
        image_list.append(img_array)
        # Save the resized image
        resized_img.save(f'excel_images/{count}.jpg')
        count += 1

    return image_list


def enhance_image(image_array, alpha=1.5, beta=30):
    enhanced_image = cv2.convertScaleAbs(image_array, alpha=alpha, beta=beta)
    return enhanced_image


def pdf_to_jpg(pdf_path, output_folder):
    # Open the PDF file
    pdf_document = fitz.open(pdf_path, )
    images = []
    # arrays = []
    jpg_folder = output_folder

    # Iterate through each page in the PDF
    if len(pdf_document) > 1:
        for page_number in range(len(pdf_document)):
            page = pdf_document[page_number]

            # Create an image of the page (DPI setting can be adjusted)
            px = page.get_pixmap(matrix=fitz.Matrix(300 / 72, 300 / 72))

            # Convert Pixmap object to Numpy array
            image_array = np.frombuffer(
                px.samples, dtype=np.uint8).reshape(px.h, px.w, px.n)

            enhanced_image = enhance_image(image_array)

            images.append(enhanced_image)

            # Save the image using OpenCV
            # jpg_path = os.path.join(output_folder, f"page_{page_number + 1}.jpg")
            # cv2.imwrite(jpg_path, cv2.cvtColor(image_array, cv2.COLOR_RGB2BGR))
    else:
        page = pdf_document[len(pdf_document)]
        px = page.get_pixmap()
        image_array = np.frombuffer(
            px.samples, dtype=np.uint8).reshape(px.h, px.w, px.n)
        images.append(image_array)

        # jpg_path = os.path.join(output_folder, "page_1.jpg")
        # cv2.imwrite(jpg_path, cv2.cvtColor(image_array, cv2.COLOR_RGB2BGR))

    # Close the PDF file
    pdf_document.close()
    # np.save(os.path.join(jpg_folder, "arrays.npy"), np.array(arrays))
    return images


def extract_years_from_dates(date_list):
    years = set()
    year_pattern = re.compile(r'\b\d{4}\b')

    current_year = datetime.datetime.today().year

    for date_str in date_list:
        year_matches = re.findall(year_pattern, date_str)
        for year in year_matches:
            year_int = int(year)
            if current_year - 5 <= year_int <= current_year:
                years.add(year)

    return list(years)


def image_to_text(OCR_path, *args):

    for arg in args:

        if not isinstance(arg, list):
            raise TypeError("Input parameter must be a list of numpy arrays.")

        else:

            pytesseract.pytesseract.tesseract_cmd = OCR_path

            final_json = {}
            master_list = []

            for _, images in enumerate(arg[64:65]):

                # img_array = img
                # image = cv2.resize(images, None, fx=0.625, fy=0.625)
                # gray_image = cv2.cvtColor(images, cv2.COLOR_BGR2GRAY)
                #
                # threshold_img = cv2.adaptiveThreshold(
                #     gray_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 81, 15)

                # Perform OCR to extract text from the image
                text = pytesseract.image_to_string(
                    images, config=r'--psm 4')
                # print("Text extracted from OCR:\n", text)

                # Split the text into lines
                lines = text.split('\n')
                # print(lines)

                date_pattern = r'\b(?:\d{1,2}[-/]\d{1,2}[-/]\d{2,4}|' \
                               r'\d{1,2} (?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]* \d{2,4}|' \
                               r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]* \d{1,2})\b'  # For Months separation """

                date_pattern2 = r'\b(?:\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|' \
                                r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-zA-Z.,-]*[\s-]?\d{1,2}?[,\s-]?[\s]?\d{4}|' \
                                r'\d{1,2}[/-]\d{4}|\d{4})\b'

                dates = []
                for values in lines:
                    if re.findall(date_pattern, values):
                        dates.append(values)

                # dates = re.findall(date_pattern, text)
                print("Dates extracted:", dates)
                supp_dates = []
                for date in dates:
                    if bool(re.search(r'[a-zA-Z]', date)) and bool(re.search(r'\d', date)):
                        supp_dates.append(date)

                print("Dates in supp_dates:", supp_dates)

                duration_keywords = ['months', 'days', 'years']

                # For keeping value type as is for the numerals
                def keep_int_or_float(value):
                    try:
                        if isinstance(int(value), int):
                            return int(value)
                    except:
                        try:
                            if isinstance(float(value), float):
                                return float(value)
                        except:
                            return None

                ## Regex pattern to have int or float values or if they have comma in between or they have parenthesis at the end
                num_pattern = r'[-+]?\d{1,5}(?:,?\d{3})*(?:\.\d+)?|\(\s*[-+]?\d{1,3}(?:,?\d{3})*(?:\.\d+)?\s*\)'
                pattern = r'[,:()\-]'
                year_list = None
                unstruct_data = []
                prev_lookup = []
                i = 0
                for word_list in lines:
                    # print(word_list)
                    num_list = re.findall(num_pattern, word_list)
                    # print("Initial num list:", num_list)
                    num_list = [num.replace(',', '') for num in num_list]
                    #                     print("Num replace list after comma:", num_list)
                    num_list = [
                        str(float(number.replace('(', '').replace(')', '')) * (-1)) if '(' in number else str(number)
                        for number in num_list]
                    #                     print("Num replace list after (): ", num_list)
                    word_list = re.sub(pattern, '', word_list)
                    word_list = word_list.split()
                    label = "_".join(
                        [word for word in word_list if word.isalpha()]).lower()
                    # print(num_list)
                    if len(num_list) != 0:
                        num_list = list(map(keep_int_or_float, num_list))
                        #                         print("Num list after mapping", num_list)

                        if all(isinstance(x, int) for x in num_list):
                            if all((x >= datetime.datetime.today().year - 6 and x <= datetime.datetime.today().year) for
                                   x in num_list):
                                year_list = num_list
                                print("Year list inside the loop:", year_list)
                        elif year_list is None:
                            dates2 = []
                            for values in lines:
                                if re.findall(date_pattern2, values):
                                    dates2.append(values)
                            year_list = extract_years_from_dates(dates2)
                            print("Year list inside the loop in else condition:", year_list)

                        unstruct_data.append({label: num_list})
                        # print(f"Unstruct Data {i+1} after year list if-else", unstruct_data)
                        # i+=1

                    for duration in duration_keywords:
                        if duration in [item.lower() for item in word_list]:
                            has_number = any(item.isdigit() for item in word_list)
                            if has_number:
                                nl = "_".join(word_list).lower()
                                prev_lookup.append(nl)
                            else:
                                prev_lookup.append(label)


                month_num = []
                for looks in prev_lookup:
                    for lk in [look for look in looks.split('_')]:
                        try:
                            month_num.append(w2n.word_to_num(lk))
                        except:
                            pass

                resultant = []
                # if len(month_num) != 0:
                #     month_num.sort()
                #     for year in set(year_list):
                #         for month in set(supp_dates):
                #             datestring = month + ', ' + str(year)
                #             date_obj = datetime.datetime.strptime(
                #                 datestring, '%B %d, %Y')
                #             for nums in month_num:
                #                 if nums == 3:
                #                     prev_date = date_obj - \
                #                                 datetime.timedelta(nums * 30)
                #                     prev_date = prev_date.strftime("%b, %Y")
                #                     resultant.append(prev_date + '-' + datestring)
                #
                #                 if nums == 6:
                #                     prev_date = date_obj - \
                #                                 datetime.timedelta(nums * 30)
                #                     prev_date = prev_date.strftime("%b, %Y")
                #                     resultant.extend([prev_date + '-' + datestring])
                #
                #                 if nums == 9:
                #                     prev_date = date_obj - \
                #                                 datetime.timedelta(nums * 30)
                #                     prev_date = prev_date.strftime("%b, %Y")
                #                     resultant.extend([prev_date + '-' + datestring])
                #
                #                 if nums == 12:
                #                     prev_date = date_obj - \
                #                                 datetime.timedelta(nums * 30)
                #                     prev_date = prev_date.strftime("%b, %Y")
                #                     resultant.extend([prev_date + '-' + datestring])

                resultant.sort()
                print("Year_list before reversal:", year_list)
                # year_list.reverse()
                print("Final Year List:", year_list)
                # if len(month_num) != 0 and len(resultant) > 0:
                #     for dictionaries in unstruct_data:
                #         for key in dictionaries.keys():
                #             if key != '':
                #                 length = len(dictionaries[key])
                #                 while length < len(resultant):
                #                     rem = len(resultant) - length
                #                     for i in range(rem):
                #                         dictionaries[key].insert(i, 0)
                #                     length = len(dictionaries[key])
                #
                #     for idx in range(len(resultant)):
                #         temp_list = []
                #         for dic in unstruct_data:
                #             for key in dic.keys():
                #                 if key != '':
                #                     temp_list.append({key: dic[key][idx]})
                #         final_json[resultant[idx]] = temp_list
                print("Unstruct data before 0's replacements:", unstruct_data)
                if year_list != None:
                    for dictionaries in unstruct_data:
                        for key in dictionaries.keys():
                            if key != '':
                                length = len(dictionaries[key])
                                while length < len(year_list):
                                    rem = len(year_list) - length
                                    for i in range(rem):
                                        dictionaries[key].insert(i, 0)
                                    length = len(dictionaries[key])

                    print("Unstruct_data after the 0's replacement:", unstruct_data)
                    for idx in range(len(year_list)):
                        temp_list = []
                        for dic in unstruct_data:
                            for key in dic.keys():
                                if key != '':
                                    temp_list.append({key: dic[key][idx]})
                        final_json[year_list[idx]] = temp_list

                else:
                    for dictionaries in unstruct_data:
                        for key in dictionaries.keys():
                            if key != '':
                                length = len(dictionaries[key])
                                while length < len(supp_dates):
                                    rem = len(supp_dates) - length
                                    for i in range(rem):
                                        dictionaries[key].insert(i, 0)
                                    length = len(dictionaries[key])

                    for idx in range(len(supp_dates)):
                        temp_list = []
                        for dic in unstruct_data:
                            for key in dic.keys():
                                if key != '':
                                    temp_list.append({key: dic[key][idx]})
                        final_json[supp_dates[idx]] = temp_list

                # key_list = final_json.keys()
                # if ((len(master_list)<=len(year_list)) or (len(master_list)<=len(resultant)) or (len(master_list)<=len(supp_dates))) and (len(master_list)!=0):
                #     for my_dict in master_list:
                #         for key in my_dict.keys():
                #             if key in key_list:
                #                 my_dict[key] = final_json[key]
                # else:
                #         master_list.append(final_json)

                master_list.append(final_json)

                # print(final_json)

                # image = cv2.resize(images, None, fx=0.48, fy=0.35)
                # gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
                # threshold_img = cv2.adaptiveThreshold(
                #     gray_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 81, 15)
                # cv2.imshow('Grayscale', threshold_img)
                # cv2.imwrite('bhel_bs.jpeg', threshold_img)
                # cv2.waitKey(0)
                #
                # # Window shown waits for any key pressing event
                # cv2.destroyAllWindows()

        print(master_list)
        return master_list


#
pdf_file = "input_pdf/zomato.pdf"  # Replace with your PDF file path
path_ocr = r"C:\Program Files\Tesseract-OCR\tesseract.exe"  # Replace with your path
output_folder = r"C:\Users\DELL\Desktop\pdf_extractor\jpg_folder"
img = pdf_to_jpg(pdf_file, output_folder)

line_items = (image_to_text(path_ocr, img))

with open('line_items_ups_cf.json', 'w') as json_file:
    json.dump(line_items, json_file)
